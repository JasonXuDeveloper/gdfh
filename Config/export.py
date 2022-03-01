#!/usr/bin/python
# -*- coding: UTF-8 -*-
import openpyxl
import sys
import json

# excel表格转json文件
def excel_to_json(excel_file, json_f_name):
    jd = []
    heads = []
    book = openpyxl.load_workbook(excel_file)
    sheet = book[u'Sheet1']
    
    max_row = sheet.max_row
    max_column = sheet.max_column
    # 解析表头
    for column in range(max_column):
        heads.append(sheet.cell(1, column + 1).value)
    # 遍历每一行
    for row in range(max_row):
        if row < 2:
        	# 前两行跳过
            continue
        one_line = {}
        # 遍历一行中的每一个单元格
        for column in range(max_column): 
            k = heads[column]
            if not k:
                continue
            v = sheet.cell(row + 1, column + 1).value
            one_line[k] = v
        jd.append(one_line)
    book.close()
    # 将json保存为文件
    save_json_file(jd, json_f_name)

# 将json保存为文件
def save_json_file(jd, json_f_name):
    f = open(json_f_name, 'w', encoding='utf-8')
    txt = json.dumps(jd, indent=2, ensure_ascii=False)
    f.write(txt)
    f.close()

e = sys.argv[1]
f = sys.argv[2]
f2 = sys.argv[3]
excel_to_json(e,f)
excel_to_json(e,f2)
print("成功将 表："+str(e)+" 转换为 配置文件："+str(f2))
